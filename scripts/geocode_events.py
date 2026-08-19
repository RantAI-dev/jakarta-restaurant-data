#!/usr/bin/env python3
"""Geocode the 'lokasi' column of 202607 Jumlah Pengunjung Event.xlsx
using the free OpenStreetMap Nominatim API (no Google Maps token needed).

Adds columns: alamat, lat, lon, geocode_match, geocode_query.
Caches results in scripts/geocode_cache.json so reruns are ~instant.
"""
import json, os, re, sys, time, urllib.parse, urllib.request

SRC = "202607 Jumlah Pengunjung Event.xlsx"
OUT = "202607 Jumlah Pengunjung Event - GEOCODED.xlsx"
CACHE = "scripts/geocode_cache.json"
UA = "DisparEventGeocode/1.0 (kleopasevan@gmail.com)"

import openpyxl

cache = {}
if os.path.exists(CACHE):
    with open(CACHE) as f:
        cache = json.load(f)

def clean(loc):
    if not loc:
        return ""
    s = str(loc).replace("\n", ", ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s).strip(" ,")
    return s

# hints that the venue is NOT in Jakarta -> don't append ", Jakarta"
FOREIGN = re.compile(r"\b(BALI|NUSA DUA|THAILAND|PATTAYA|DENMARK|COPENHAGEN|KOREA|"
                     r"NONSAN|JAPAN|SINGAPORE|MALAYSIA|BANDUNG|SURABAYA|YOGYA|BOGOR|"
                     r"BEKASI|DEPOK|TANGERANG|BANTEN|SEMARANG|MEDAN|BALIKPAPAN|LUAR NEGERI)\b",
                     re.I)

def nominatim(q):
    url = ("https://nominatim.openstreetmap.org/search?"
           + urllib.parse.urlencode({"q": q, "format": "json", "limit": 1,
                                     "addressdetails": 1}))
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=25) as r:
        data = json.load(r)
    time.sleep(1.1)  # respect Nominatim usage policy
    return data

def geocode(loc):
    key = clean(loc).upper()
    if not key:
        return None
    if key in cache:
        return cache[key]
    base = clean(loc)
    queries = [base]
    if not FOREIGN.search(base) and not re.search(r"\bJAKARTA\b", base, re.I):
        queries.append(base + ", Jakarta, Indonesia")
    elif not FOREIGN.search(base):
        queries.append(base + ", Indonesia")
    result = None
    for i, q in enumerate(queries):
        try:
            data = nominatim(q)
        except Exception as e:
            print(f"  ! error on {q!r}: {e}", file=sys.stderr)
            data = []
        if data:
            hit = data[0]
            result = {
                "alamat": hit.get("display_name", ""),
                "lat": hit.get("lat", ""),
                "lon": hit.get("lon", ""),
                "match": "exact" if i == 0 else "fallback-jakarta",
                "query": q,
            }
            break
    if result is None:
        result = {"alamat": "", "lat": "", "lon": "", "match": "NOT FOUND", "query": queries[-1]}
    cache[key] = result
    with open(CACHE, "w") as f:
        json.dump(cache, f, ensure_ascii=False, indent=0)
    return result

def main():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["SDI"]
    header = [c.value for c in ws[1]]
    loc_col = header.index("lokasi") + 1  # 1-based

    # collect unique locations first (for progress + dedup)
    locs = {}
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=loc_col).value
        if v and clean(v):
            locs.setdefault(clean(v).upper(), v)
    total = len(locs)
    print(f"{total} unique locations to geocode "
          f"({sum(1 for k in locs if k in cache)} already cached)")

    n = 0
    for key, sample in locs.items():
        n += 1
        r = geocode(sample)
        tag = r["match"]
        print(f"[{n}/{total}] {tag:16} {clean(sample)[:50]}")

    # write new columns
    new_cols = ["alamat", "lat", "lon", "geocode_match", "geocode_query"]
    start = ws.max_column + 1
    for j, name in enumerate(new_cols):
        ws.cell(row=1, column=start + j, value=name)
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=loc_col).value
        r = cache.get(clean(v).upper()) if v and clean(v) else None
        if not r:
            r = {"alamat": "", "lat": "", "lon": "", "match": "", "query": ""}
        vals = [r["alamat"], r["lat"], r["lon"], r["match"], r["query"]]
        for j, val in enumerate(vals):
            ws.cell(row=row, column=start + j, value=val)

    wb.save(OUT)
    found = sum(1 for r in cache.values() if r["match"] != "NOT FOUND")
    print(f"\nSaved {OUT}")
    print(f"Coverage: {found}/{total} located, "
          f"{total - found} not found.")

if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Merge crawled addresses (with sumber) into final xlsx, geo columns placed
right after 'lokasi'. Re-crawl results override the first crawl; first crawl
overrides raw OSM. Also emits a review CSV of rows a human should eyeball.
"""
import csv, json, os, re
import openpyxl

SRC = "202607 Jumlah Pengunjung Event.xlsx"
OUT = "202607 Jumlah Pengunjung Event - GEOCODED.xlsx"
REVIEW = "202607 Event Geocode - PERLU REVIEW.csv"
GEO = ["alamat", "kota", "lat", "lon", "sumber", "confidence", "review_status", "catatan"]

def clean(s):
    return re.sub(r"\s+", " ", str(s).replace("\n", ", ").replace("\r", " ")).strip(" ,")

def load(path):
    if not os.path.exists(path):
        return {}
    return {r["lokasi"].strip().upper(): r for r in json.load(open(path))}

crawl = load("scripts/crawl_results.json")
recrawl = load("scripts/recrawl_results.json")
cache = json.load(open("scripts/geocode_cache.json"))

def from_crawl(r):
    return {"alamat": r["alamat"], "kota": r["kota"], "lat": r["lat"], "lon": r["lon"],
            "sumber": r["sumber"], "confidence": r["confidence"],
            "review_status": r["review_status"], "catatan": r["catatan"]}

def record(loc):
    key = clean(loc).upper()
    # prefer re-crawl, unless it downgraded a previously-good crawl to not_found
    if key in recrawl and recrawl[key]["confidence"] != "not_found":
        return from_crawl(recrawl[key])
    if key in crawl and crawl[key]["confidence"] != "not_found":
        return from_crawl(crawl[key])
    if key in recrawl:
        return from_crawl(recrawl[key])
    if key in crawl:
        return from_crawl(crawl[key])
    c = cache.get(key)
    if c and c["match"] == "exact":
        return {"alamat": c["alamat"], "kota": "", "lat": c["lat"], "lon": c["lon"],
                "sumber": "OpenStreetMap / Nominatim", "confidence": "medium",
                "review_status": "auto-osm", "catatan": ""}
    return {k: "" for k in GEO} | {"confidence": "not_found", "review_status": "not_found"}

wb = openpyxl.load_workbook(SRC)
ws = wb["SDI"]
orig = [c.value for c in ws[1]]
while orig and orig[-1] in (None, ""):  # drop trailing empty columns
    orig.pop()
loc_i = orig.index("lokasi")
# new header order: cols up to & incl lokasi, then GEO, then the rest
new_header = orig[:loc_i + 1] + GEO + orig[loc_i + 1:]
old_rows = list(ws.iter_rows(min_row=2, values_only=True))

# build fresh sheet
nwb = openpyxl.Workbook()
nws = nwb.active
nws.title = "SDI"
nws.append(new_header)

review_rows, seen = [], set()
for row in old_rows:
    v = row[loc_i]
    if v is None or not clean(v):
        # keep blank/spacer rows as-is (padded to new width)
        r = {k: "" for k in GEO}
        out = list(row[:loc_i + 1]) + [r[k] for k in GEO] + list(row[loc_i + 1:len(orig)])
        nws.append(out)
        continue
    r = record(v)
    out = list(row[:loc_i + 1]) + [r[k] for k in GEO] + list(row[loc_i + 1:len(orig)])
    nws.append(out)
    if r["review_status"] in ("needs_review", "rejected", "not_found"):
        k = clean(v).upper()
        if k not in seen:
            seen.add(k)
            review_rows.append({"lokasi": clean(v), **r})

nwb.save(OUT)
with open(REVIEW, "w", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["lokasi"] + GEO)
    w.writeheader()
    for r in sorted(review_rows, key=lambda x: x["review_status"]):
        w.writerow(r)

from collections import Counter
tier = Counter()
for row in old_rows:
    v = row[loc_i]
    if v is None or not clean(v):
        continue
    tier[record(v)["review_status"]] += 1
tot = sum(tier.values())
print(f"Saved {OUT}  (geo columns now after 'lokasi')")
print(f"Review file: {REVIEW} ({len(review_rows)} unique to eyeball)")
print(f"\nRow-level ({tot} events):")
for k, n in tier.most_common():
    print(f"  {k:14} {n:4}  {100*n/tot:3.0f}%")

#!/usr/bin/env python3
"""Second pass: fill NOT-FOUND locations using Photon (free, fuzzy POI search),
with guardrails so we don't accept far-away false matches.

Reads/updates scripts/geocode_cache.json, then rewrites the GEOCODED xlsx.
"""
import json, math, os, re, sys, time, urllib.parse, urllib.request

SRC = "202607 Jumlah Pengunjung Event.xlsx"
OUT = "202607 Jumlah Pengunjung Event - GEOCODED.xlsx"
CACHE = "scripts/geocode_cache.json"
JKT = (-6.2088, 106.8456)     # Monas-ish center
MAX_KM = 70                   # accept ID results within this of Jakarta

import openpyxl

FOREIGN = re.compile(r"\b(BALI|NUSA DUA|THAILAND|PATTAYA|DENMARK|COPENHAGEN|KOREA|"
                     r"NONSAN|JAPAN|SINGAPORE|MALAYSIA|BANDUNG|SURABAYA|YOGYA|BOGOR|"
                     r"BEKASI|DEPOK|TANGERANG|BANTEN|SEMARANG|MEDAN|BALIKPAPAN|LUAR NEGERI)\b",
                     re.I)

def clean(loc):
    s = str(loc).replace("\n", ", ").replace("\r", " ")
    return re.sub(r"\s+", " ", s).strip(" ,")

def haversine(a, b):
    R = 6371
    dlat = math.radians(b[0]-a[0]); dlon = math.radians(b[1]-a[1])
    x = (math.sin(dlat/2)**2 + math.cos(math.radians(a[0]))*math.cos(math.radians(b[0]))
         * math.sin(dlon/2)**2)
    return 2*R*math.asin(math.sqrt(x))

def photon(q):
    url = "https://photon.komoot.io/api/?" + urllib.parse.urlencode(
        {"q": q, "limit": 5, "lat": JKT[0], "lon": JKT[1]})
    req = urllib.request.Request(url, headers={"User-Agent": "DisparEventGeocode/1.0"})
    with urllib.request.urlopen(req, timeout=25) as r:
        data = json.load(r)
    time.sleep(0.6)
    return data.get("features", [])

def assemble(p):
    parts = [p.get("name"), p.get("street"), p.get("district"),
             p.get("city"), p.get("state"), p.get("country")]
    seen, out = set(), []
    for x in parts:
        if x and x not in seen:
            seen.add(x); out.append(x)
    return ", ".join(out)

def main():
    cache = json.load(open(CACHE))
    todo = [(k, v) for k, v in cache.items() if v["match"] == "NOT FOUND"]
    print(f"{len(todo)} NOT FOUND -> trying Photon")
    filled = 0
    for i, (key, v) in enumerate(todo, 1):
        base = clean(key)
        foreign = bool(FOREIGN.search(base))
        try:
            feats = photon(base)
        except Exception as e:
            print(f"  ! {base[:40]}: {e}", file=sys.stderr); feats = []
        chosen = None
        for f in feats:
            p = f["properties"]; g = f["geometry"]["coordinates"]  # [lon,lat]
            latlon = (g[1], g[0])
            if foreign:
                chosen = (p, latlon); break
            if p.get("countrycode") == "ID" and haversine(JKT, latlon) <= MAX_KM:
                chosen = (p, latlon); break
        if chosen:
            p, latlon = chosen
            v.update(alamat=assemble(p), lat=str(latlon[0]), lon=str(latlon[1]),
                     match="photon (review)",
                     query=f"photon:{base[:40]} -> {p.get('name','')}")
            filled += 1
            tag = "OK "
        else:
            tag = "no "
        print(f"[{i}/{len(todo)}] {tag}{base[:48]}")
    json.dump(cache, open(CACHE, "w"), ensure_ascii=False, indent=0)

    # rewrite xlsx
    wb = openpyxl.load_workbook(SRC)
    ws = wb["SDI"]
    header = [c.value for c in ws[1]]
    loc_col = header.index("lokasi") + 1
    new_cols = ["alamat", "lat", "lon", "geocode_match", "geocode_query"]
    start = ws.max_column + 1
    for j, name in enumerate(new_cols):
        ws.cell(row=1, column=start+j, value=name)
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=loc_col).value
        r = cache.get(clean(val).upper()) if val and clean(val) else None
        r = r or {"alamat":"","lat":"","lon":"","match":"","query":""}
        for j, x in enumerate([r["alamat"], r["lat"], r["lon"], r["match"], r["query"]]):
            ws.cell(row=row, column=start+j, value=x)
    wb.save(OUT)

    from collections import Counter
    print("\n", Counter(v["match"] for v in cache.values()))
    found = sum(1 for v in cache.values() if v["match"] != "NOT FOUND")
    print(f"Coverage now: {found}/{len(cache)}  (+{filled} from Photon)")

if __name__ == "__main__":
    main()

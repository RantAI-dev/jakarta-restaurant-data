"""Sumber berkas lokal: TSV, CSV, XLSX.

Dipakai untuk data sekunder yang tidak ada di SDI — hasil crawl, kompilasi
manual, dan lampiran dari dinas (data GCI/GPCI, halal, event, kontak DTW).
"""

from __future__ import annotations

import csv
import os
from typing import Any, Iterator

EKSTENSI = {".tsv", ".csv", ".xlsx", ".json"}

# Berkas kerja/cadangan yang tidak boleh ikut masuk lake.
POLA_ABAIKAN = ("~$", ".bak.", "-KERJA", "-SWEEP", "PERLU REVIEW")


def discover(root: str) -> list[str]:
    """Cari berkas data yang layak diambil, rekursif, terurut stabil."""
    hasil: list[str] = []
    for dirpath, dirnames, filenames in os.walk(root):
        dirnames[:] = [
            d for d in dirnames if not d.startswith(".") and d != "node_modules"
        ]
        for fn in filenames:
            if os.path.splitext(fn)[1].lower() not in EKSTENSI:
                continue
            if any(p in fn for p in POLA_ABAIKAN):
                continue
            hasil.append(os.path.join(dirpath, fn))
    return sorted(hasil)


def _sniff_delimiter(path: str) -> str:
    if path.lower().endswith(".tsv"):
        return "\t"
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        sample = f.read(8192)
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def read_delimited(path: str) -> Iterator[dict[str, Any]]:
    delim = _sniff_delimiter(path)
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f, delimiter=delim)
        header = reader.fieldnames or []
        for row in reader:
            # Buang baris kosong total dan baris header yang terulang di tengah.
            values = [v for v in row.values() if v not in (None, "")]
            if not values:
                continue
            if list(row.values())[: len(header)] == header:
                continue
            row.pop(None, None)  # kolom berlebih dari baris rusak
            yield {k: v for k, v in row.items() if k is not None}


def read_xlsx(path: str) -> Iterator[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        rows = ws.iter_rows(values_only=True)
        header: list[str] = []
        for raw in rows:
            if raw is None:
                continue
            if not header:
                if all(c is None for c in raw):
                    continue
                header = [
                    str(c).strip() if c is not None else f"kolom_{i}"
                    for i, c in enumerate(raw)
                ]
                continue
            if all(c is None for c in raw):
                continue
            yield {header[i]: raw[i] for i in range(min(len(header), len(raw)))}
    finally:
        wb.close()


def read_json(path: str) -> Iterator[dict[str, Any]]:
    """Array JSON berisi objek datar → baris. Nilai nested di-serialize di Bronze."""
    import json

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        # Kalau objek tunggal membungkus array (mis. {"data": [...]}), pakai array
        # pertama yang ditemukan; kalau tidak, perlakukan sebagai satu baris.
        arr = next((v for v in data.values() if isinstance(v, list)), None)
        data = arr if arr is not None else [data]
    if not isinstance(data, list):
        return
    for item in data:
        if isinstance(item, dict):
            yield item


def read_file(path: str) -> Iterator[dict[str, Any]]:
    low = path.lower()
    if low.endswith(".xlsx"):
        return read_xlsx(path)
    if low.endswith(".json"):
        return read_json(path)
    return read_delimited(path)
